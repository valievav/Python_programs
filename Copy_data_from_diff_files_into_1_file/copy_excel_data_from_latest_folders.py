import datetime
import os
import re

import openpyxl
import logging


def find_all_latest_folders(cwd: str, exclude_folders: list, priority_keyword: str)-> list:
    """
    Returns ALL most recent folders in a directory.\n
    The general flow of the function is next: find folders that start with 'YYYYMM' -> compare dates for such folders
    withing the parent directory -> detect folder with the latest date.\n
    If found several YYYYMM folders with the same date - use priority keyword to determine which folder
    should be considered the most recent one.\n
    Returns all_latest_folders_data[[folder_date, folder_name],...]
    """

    def exclude_folder(current_folder: str, folders_to_exclude: list)-> bool:
        """
        Returns True flag if folder is in the exclusion list.
        """

        exclude = False
        for folder in folders_to_exclude:
            if folder.lower() in [folder.lower() for folder in current_folder.split(os.sep)]:
                logging.debug(f"SKIPPING  path '{os.path.join(current_folder)}' because of exclusion folder '{folder}'")
                exclude = True
        return exclude

    def find_one_latest_folder(folders: list)-> list:
        """
        Returns 1 most recent folder inside of the directory.\n
        The general flow of the function is next: find folders that start with YYYYMM, parse date
        and determine the most recent one.\n
        If found several YYYYMM folders with the same date - use priority keyword to determine which folder
        should be considered the most recent one.
        Returns latest_folder_data[[folder_date, folder_name]
        """

        def yyyymm_format_match(year_month: str, folder: str)-> bool:
            """
            Detects if value is in YYYYMM format.
            """
            year = int(year_month[:4])
            month = int(year_month[4:])
            if 0 == month > 12:
                logging.debug(f"DATE PARSING ERROR: actual MONTH - '{month}', expected MONTH in range (1-12)."
                              f"Skipping folder '{folder}'.")
                return False
            elif 2010 <= year > datetime.datetime.today().year:
                logging.debug(f"DATE PARSING ERROR: actual YEAR - '{year}', expected YEAR in range (2010-now). "
                              f"Skipping folder '{folder}'.")
                return False
            else:
                return True

        latest_folder_data = []
        for folder in (_ for _ in folders):  # folders generator

            # find YYYYMM match
            regex_match = re.match(r"^\d{6}", folder)
            if not regex_match:
                continue
            year_month = regex_match.group()
            if not yyyymm_format_match(year_month, folder):
                continue
            folder_date = datetime.datetime.strptime(year_month, "%Y%m").date()

            # determine most recent folder
            if not latest_folder_data or folder_date > latest_folder_data[0]:  # if empty or > than existing
                latest_folder_data = [folder_date, folder]
            elif folder_date == latest_folder_data[0]:  # if dates the same - use keyword
                if priority_keyword in folder:
                    logging.debug(f"Found priority keyword '{priority_keyword}' in folder '{folder}'. "
                                  f"Previous folder '{latest_folder_data[1]}' is no longer latest.")
                    latest_folder_data = [folder_date, folder]
        return latest_folder_data

    # iterate through subfolders and files
    os.chdir(cwd)
    all_latest_folders_data = []
    for active_folder, subfolders, files in os.walk('.'):

        if exclude_folder(active_folder, exclude_folders):
            continue

        latest_folder_data = find_one_latest_folder(subfolders)
        if latest_folder_data:
            latest_date = latest_folder_data[0]
            latest_folder = latest_folder_data[1]
            latest_folder_rel_path = os.path.relpath(os.path.join(active_folder, latest_folder), cwd)

            all_latest_folders_data.append([latest_date, latest_folder_rel_path])
            print(f"    DETECTED latest folder in '{active_folder}' -> latest DATE - '{latest_date}', "
                  f"latest FOLDER - '{latest_folder}' out of {subfolders}")
    return all_latest_folders_data


def copy_data_into_master_file(cwd: str, target_folders_data: list, file_keywords: list, sheet_keywords: list,
                               master_file_path: str, master_file: str, recreate_master_file=True)-> None:
    """
    Finds files that match keywords and copies data from them into Master file.\n
    The general flow of the function is next: find files that contain keyword in name in the provided directory ->
    find sheet inside the file, that contain keyword in sheet name -> copy ALL data from matched sheet
    into newly created Master file.
    """

    def keyword_match(elems, keywords: list)-> dict:
        """
        Finds keyword matches for any list/str.\n
        Returns dictionary {keyword:matched_element}.
        """

        matches = {}
        if isinstance(elems, str):
            elems = [elems]  # list with 1 element - for file match case

        [matches.setdefault(keyword, elem) for keyword in keywords for elem in elems if keyword.lower() in elem.lower()]
        return matches

    def record_data_to_sheet(parent_folder: str, current_folder_date: datetime, master_file_abs_path: str,
                             from_file_abs_path: str, sheet_name: str, data_to_record: iter,
                             recreate_file: bool)-> None:
        """
        Reads data from file and records it into Master Excel file.\n
        Uses recreate_file flag to determine whether to recreate Master file or update it:
        new file created only on first iteration and updated on all next iterations.
        """

        # remove old file only on the start of the program
        if recreate_file and os.path.isfile(master_file_abs_path):
            os.unlink(master_file_abs_path)
            logging.debug(f"REMOVED old result file '{master_file_abs_path}'")

        try:
            master_file_wb = openpyxl.load_workbook(master_file_abs_path)
        except FileNotFoundError:
            master_file_wb = openpyxl.Workbook()
            logging.debug(f"CREATED new result file")

        # create sheet if not exists
        if sheet_name not in master_file_wb.sheetnames:
            master_file_wb.create_sheet(sheet_name, 0)
            logging.debug(f"CREATED new sheet - '{sheet_name}'")
        master_file_sheet = master_file_wb[sheet_name]

        # record data
        metadata = (from_file_abs_path, parent_folder, current_folder_date)  # casted to tuple for further concatenation
        for row in data_to_record:
            empty_row = all(value is None for value in row)
            if not empty_row:
                record_line = metadata + row
                master_file_sheet.append(record_line)  # append works only with iterables
        master_file_sheet.append([""])  # empty row to separate the data from different files
        print(f"    RECORDED data from FILE '{from_file_abs_path}' into SHEET '{sheet_name}'")

        master_file_wb.save(master_file_abs_path)

    # iterate through target folders to find matched file, sheet and copy data
    os.chdir(cwd)
    for target_folder_date, target_folder in (_ for _ in target_folders_data):  # target_folders_data generator
        for active_folder, subfolders, files in os.walk(target_folder):
            for file in (_ for _ in files):  # files generator

                # match file
                matched_file = keyword_match(elems=file, keywords=file_keywords)
                if matched_file:
                    file_abs_path = os.path.join(os.getcwd(), active_folder, file)
                    wb = openpyxl.load_workbook(file_abs_path, read_only=True)

                    # match sheets
                    sheets = wb.sheetnames
                    matched_sheets = keyword_match(elems=sheets, keywords=sheet_keywords)

                    for matched_keyword, matched_sheet in matched_sheets.items():

                        # read data from sheet
                        row_data = wb[matched_sheet].values

                        # record data to result file sheet
                        parent_folder_name = target_folder.split(os.sep)[0]
                        master_file_abs_path = os.path.join(master_file_path, master_file)
                        record_data_to_sheet(parent_folder=parent_folder_name, current_folder_date=target_folder_date,
                                             master_file_abs_path=master_file_abs_path, from_file_abs_path=file_abs_path,
                                             sheet_name=matched_keyword, data_to_record=row_data,
                                             recreate_file=recreate_master_file)
                        recreate_master_file = False  # no dot recreate file (add data from processed files)


def main():
    search_path = r"D:\PYTHON Practice\Consolidate function folder"
    file_keywords = ["data"]
    sheet_keywords = ["detailed", "general"]
    skip_folders = ["Skipper"]
    priority_keyword = "redeliver"

    destination_path = r"D:\PYTHON Practice\Consolidate function folder - result"
    destination_file = "Consolidation results.xlsx"

    print("Process started...")
    start = datetime.datetime.now()
    logging.basicConfig(level=logging.CRITICAL, format=' %(asctime)s - %(levelname)s - %(message)s')
    latest_folders_data = find_all_latest_folders(cwd=search_path,
                                                  exclude_folders=skip_folders,
                                                  priority_keyword=priority_keyword,
                                                  )
    copy_data_into_master_file(cwd=search_path,
                               target_folders_data=latest_folders_data,
                               file_keywords=file_keywords,
                               sheet_keywords=sheet_keywords,
                               master_file_path=destination_path,
                               master_file=destination_file,
                               )

    finish = datetime.datetime.now()
    print(f"Process finished. Time spent {finish - start}")


if __name__ == "__main__":
    main()

