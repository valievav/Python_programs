import collections
import datetime
import logging
import os

import openpyxl
import send2trash
from openpyxl.utils.cell import get_column_letter


def find_keywords_in_files(search_keywords: list,
                           result_file: str,
                           directory: str = None,
                           exclusion_folders: list = None,
                           file_name_keywords: list = None,
                           sheet_name_keywords: list = None):
    """
    Finds excel files that contain certain keyword values.\n
    General flow: iterate through all files from the specified directory including subfolders,
    find all files OR files with specific keyword in name and run search in all spreadsheets
    OR spreadsheets with specific keyword in name.\n
    Result of the search is displayed in stack trace as well as saved into the result file.\n
    """

    def search_keywords_in_sheet(wb: openpyxl,
                                 sheet_name: str,
                                 search_keywords: list,
                                 file_abs_path: str):

        Result = collections.namedtuple("Result", "keyword, cell_value, sheet, coordinates, file_abs_path")
        sheet_results = []
        search_keywords = [keyword.lower() for keyword in search_keywords]  # preparing values for comparison

        # find keywords in cells (iterating over cells first and then keywords for faster run)
        cell_gen = (cell for row in wb[sheet_name].iter_rows() for cell in row if cell.value)
        for cell in cell_gen:
            for keyword in search_keywords:
                if keyword in str(cell.value).lower():
                    result_cell = Result(keyword, cell.value, sheet_name,
                                         f'{get_column_letter(cell.column)}:{cell.row}', file_abs_path)
                    sheet_results.append(result_cell)

                    # print cell results
                    print(f"    >>> FOUND RESULTS: ", end="")
                    for n in range(len(Result._fields)):
                        atr = Result._fields[n]
                        atr_value = getattr(result_cell, atr)
                        if n == len(Result._fields)-1:
                            print(f"{atr} - '{atr_value}' ")
                        else:
                            print(f"{atr} - '{atr_value}', ", end="")

        return sheet_results

    def record_results_into_file(results: list,
                                 result_file: str,
                                 recreate_file: bool):

        # create new file on each run
        if recreate_file:
            send2trash.send2trash(result_file)
            logging.debug("Removed previous result file")

        # update file on the same run
        try:
            with open(result_file, "a") as file:
                file.writelines(f"{results}\n")
        except Exception as err:
            logging.exception(f"Occurred exception {err} when trying to open file {result_file}.\n"
                              f"Please fix the issue!")

    def search_keywords_in_file(file_abs_path: str,
                                search_keywords: list,
                                sheet_name_keywords: list = None):

        try:
            wb = openpyxl.load_workbook(file_abs_path, read_only=True)
        except Exception as err:
            print(f"{err} for file '{file}'")
        else:
            # search keyword in all sheets or specific sheets
            if sheet_name_keywords:
                search_sheets = [sheet for sheet in wb.sheetnames for keyword in sheet_name_keywords
                                 if keyword.lower() in sheet.lower()]
            else:
                search_sheets = wb.sheetnames

            file_results = []
            if search_sheets:
                for sheet in search_sheets:
                    sheet_results = search_keywords_in_sheet(wb, sheet, search_keywords, file_abs_path)
                    file_results.append(sheet_results)

            return file_results

    print("Process started...")
    start = datetime.datetime.now()

    if directory:
        os.chdir(directory)

    prev_excluded_folder = None
    recreate_result_file = True

    # search for files in all subfolders
    for active_folder, subfolders, files in os.walk('.'):

        # skip subfolders if parent folder was excluded (prevents searching in depth to save time)
        if prev_excluded_folder and prev_excluded_folder in active_folder:
            continue

        # skip exclusion folders
        exclusion_folder = [folder for folder in exclusion_folders if folder in active_folder.split(os.sep)]
        if exclusion_folders and exclusion_folder:
            logging.debug(f"Excluded '{active_folder}'and it's subfolders from search as it contains "
                          f"exclusion folder '{exclusion_folder}'")
            prev_excluded_folder = active_folder
            continue

        for file in files:
            if file.endswith('xlsx'):
                file_abs_path = os.path.abspath(os.path.join(active_folder, file))  # abs path to open file

                # search in all files or ones with keyword in name
                logging.debug(f"Processing {file_abs_path} ...")

                # search in file if no keywords or keywords match
                if not file_name_keywords or [file for keyword in file_name_keywords if keyword.lower() in file.lower()]:
                    file_results = search_keywords_in_file(file_abs_path, search_keywords, sheet_name_keywords)
                    record_results_into_file(file_results, result_file, recreate_result_file)
                    recreate_result_file = False

    finish = datetime.datetime.now()
    time_spent = finish - start
    print(f"Data recorded into file '{result_file}'.\nProcess finished. Time spent {time_spent}. ")


if __name__ == "__main__":

    file_name_keywords = ["file"]
    search_keywords = ["Castle", "Final Fantasy"]
    sheet_name_keywords = ["sheet"]
    directory = r"D:\PYTHON Practice\Top Secret"
    exclusion_folders = ["Secret Secret"]
    result_file = "Result_file.txt"

    logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
    find_keywords_in_files(file_name_keywords=file_name_keywords,
                           sheet_name_keywords=sheet_name_keywords,
                           search_keywords=search_keywords,
                           directory=directory,
                           result_file=result_file,
                           exclusion_folders=exclusion_folders)


