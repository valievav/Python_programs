import openpyxl
import os
import datetime
from openpyxl.utils.cell import get_column_letter
from operator import itemgetter


def generate_schedule(file_name, schedule_data, iteration_days, iteration_number, date_format, no_schedule_days,
                      tasks_dates_sheet, final_schedule_sheet, never_executed_tasks_sheet, final_file_name, cwd=None):
    """
    Generates task schedule based on the task-date input and records results into Excel workbook.\n
    During this process 3 data structures are created:\n
    - all_tasks dictionary - tasks, last run date and next iteration run dates - schedule view separately for tasks;\n
    - tasks_without_date list - tasks that weren't scheduled before - requires attention (need to schedule 1st run);\n
    - tasks_schedule list - sorted schedule for all tasks and dates - comprehensive schedule ready to be used.\n
    Data from all 3 structures is recorded to a separate excel spreadsheet.\n
    :param file_name: .xlsx file
    :param schedule_data: dictionary
    :param iteration_days: datetime
    :param iteration_number: int
    :param date_format: str
    :param no_schedule_days: tuple with int
    :param tasks_dates_sheet: str
    :param final_schedule_sheet: str
    :param never_executed_tasks_sheet: str
    :param final_file_name: .xlsx file
    :param cwd: valid path
    :return:
    """

    def re_create_sheet(wb, sheet_name):
        """
        Creates new sheet or recreates anew if it already exists.\n
        :param wb: obj
        :param sheet_name: str
        :return:
        """
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        wb.create_sheet(sheet_name)

    def set_column_width(sheet):
        """
        Finds max cell length of the sheet and sets such width for the entire sheet.\n
        The purpose is to be able to see data without manually widening columns.\n
        :param sheet: obj
        :return:
        """
        if sheet.max_column == 1 and sheet.max_row == 1:  # handle cases with only 1 cell
            max_len = len(str(sheet["A1"].value))
        else:
            max_len = max(len(str(cell)) for row in sheet.values for cell in row)  # finds max value from all sheet cells
        for i in range(1, sheet.max_column+1):
            sheet.column_dimensions[get_column_letter(i)].width = max_len

    # switch to cwd if passed as an argument
    if cwd:
        if not os.path.isdir(cwd):
            os.mkdir(cwd)
        os.chdir(cwd)

    # read data from excel
    sheet_name = schedule_data["sheet_name"]
    tasks_col = schedule_data["tasks_col"]
    date_col = schedule_data["date_col"]

    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]

    # create dictionary with tasks and last run dates
    all_tasks = {}
    max_row = sheet.max_row

    for i in range(1, max_row+1):
        task = sheet.cell(row=i, column=tasks_col).value
        date = sheet.cell(row=i, column=date_col).value
        if isinstance(date, str):  # skip header or invalid date values
            print(f"Skipping row with invalid date value = '{date}'")
            continue
        if task not in all_tasks.keys():  # record unique tasks
            all_tasks.setdefault(task, [date])  # record date (including None)
        if not all_tasks[task][0]:
            all_tasks[task] = [date]  # record current date if previous date is None (None-Date case)
        else:
            if date:  # don't overwrite with None date (Date-None case)
                if date > all_tasks[task][0]:  # overwrite with most recent date
                    all_tasks[task] = [date]

    # calculate next execution day and separate tasks without date
    tasks_without_date = []

    for task in all_tasks.keys():
        if not all_tasks[task][0]:  # catch tasks with None date
            tasks_without_date.append(task)  # record into separate list
        else:
            for i in range(iteration_number-1):
                date = all_tasks[task][i]
                next_date = date + iteration_days
                while next_date.weekday() in no_schedule_days:
                    next_date += datetime.timedelta(1)
                all_tasks[task].append(next_date)

    # remove tasks without date from common scope since they can't be scheduled
    for task in tasks_without_date:
        del all_tasks[task]

    # record all tasks and dates to new sheet
    re_create_sheet(wb, tasks_dates_sheet)
    tasks_schedule_sheet = wb[tasks_dates_sheet]
    row = 1

    for task in all_tasks.keys():
        tasks_schedule_sheet[f"A{row}"] = task
        for col in range(1, len(all_tasks[task])+1):
            date_cell = tasks_schedule_sheet[f"{get_column_letter(col+1)}{row}"]
            date_cell.value = all_tasks[task][col-1]  # record value
            date_cell.number_format = date_format  # set date format
        row += 1

    print(f"Recorded results into '{tasks_dates_sheet}' sheet")
    set_column_width(tasks_schedule_sheet)

    # prepare task schedule and sort records
    tasks_schedule = []

    for i in range(iteration_number):
        for task in all_tasks.keys():
            tasks_schedule.append([task, all_tasks[task][i]])
    tasks_schedule = sorted(tasks_schedule, key=itemgetter(1))

    # record schedule data to new sheet
    re_create_sheet(wb, final_schedule_sheet)
    schedule_sheet = wb[final_schedule_sheet]
    for row in range(len(tasks_schedule)):
        for col in range(len(tasks_schedule[row])):
            schedule_sheet[f"{get_column_letter(col+1)}{row+1}"] = tasks_schedule[row][col]

    print(f"Recorded results into '{final_schedule_sheet}' sheet")
    set_column_width(schedule_sheet)

    # record tasks that don't have date to new sheet
    re_create_sheet(wb, never_executed_tasks_sheet)
    not_scheduled_tasks_sheet = wb[never_executed_tasks_sheet]
    for col in range(len(tasks_without_date)):
        not_scheduled_tasks_sheet[f"A{col+1}"] = tasks_without_date[col]

    print(f"Recorded results into '{never_executed_tasks_sheet}' sheet")
    set_column_width(not_scheduled_tasks_sheet)

    # save results in separate file to avoid original file corruption
    wb.save(final_file_name)
    print(f"Results are saved in file '{file_final}'")


if __name__ == "__main__":

    working_directory = "D:\\Practice Python\\Tasks schedule generator"
    file = "Tasks_schedule.xlsx"
    file_data = {"sheet_name": "Schedule", "tasks_col": 1, "date_col": 2}
    iterations_days = datetime.timedelta(90)
    iterations_number = 4
    date_formatting = "MM/DD/YYYY"
    do_not_schedule_days = (0, 1, 6)  # do not schedule for Monday and weekend
    tasks_dates_sheet_name = "Tasks_and_execution_dates"
    final_schedule_sheet_name = "Tasks_schedule"
    never_executed_tasks_sheet_name = "Tasks_never_executed"
    file_final = "Tasks_schedule_final.xlsx"

    generate_schedule(file, file_data, iterations_days, iterations_number, date_formatting, do_not_schedule_days,
                      tasks_dates_sheet_name, final_schedule_sheet_name, never_executed_tasks_sheet_name, file_final)


