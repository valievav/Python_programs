import openpyxl
import os
import datetime
import itertools
import logging


def consolidate_excel_data(cwd: str, file_keyword: str, sheet_keyword: str, result_path: str, result_file_name: str,
                           exclude_folders: list) -> None:

    start = datetime.datetime.now()
    print("Process started...")

    # open/create result file for writing
    result_file = os.path.join(result_path, result_file_name)
    try:
        wb_result = openpyxl.load_workbook(result_file)
    except FileNotFoundError:
        wb_result = openpyxl.Workbook()

    sheet_names = itertools.count(1, 1)  # for naming result file sheets (new sheet for each file)

    # iterate through subfolders and files
    for active_folder, subfolders, files in os.walk(cwd):
        file_gen = (file for file in files)

        exclude_folder = False
        for folder in exclude_folders:
            if folder in active_folder.split(os.sep):
                print(f"    Skipping  path '{os.path.join(active_folder)}' because of the excluded folder '{folder}'")
                exclude_folder = True

        if not exclude_folder:
            for file in file_gen:
                abs_path = os.path.join(active_folder, file)

                # check for file and keyword match
                if file_keyword.lower() in file.lower():
                    wb = openpyxl.load_workbook(abs_path)

                    for sheet in wb.sheetnames:
                        if sheet_keyword.lower() in sheet.lower():

                            # get sheet data
                            sheet = wb[sheet]
                            row_data = sheet.values

                            # recreate sheet for existing wb or create new for new wb
                            sheet_name = str(next(sheet_names))
                            if sheet_name in wb_result.sheetnames:
                                wb_result.remove(wb_result[sheet_name])
                                wb_result.create_sheet(sheet_name)
                                logging.debug(f"Recreated existing sheet '{sheet_name}'")
                            else:
                                wb_result.create_sheet(sheet_name)
                                logging.debug(f"Created new sheet '{sheet_name}'")

                            result_sheet = wb_result[sheet_name]

                            # record data to result file
                            result_sheet.append([f"Data file path {abs_path}"])  # append works only with iterables
                            for row in row_data:
                                result_sheet.append(row)
                            print(f"Recorded data from file '{abs_path}' to sheet '{sheet_name}'")

                            wb_result.save(result_file)

    finish = datetime.datetime.now()
    print(f"Process finished. Time spent {finish - start}")


if __name__ == "__main__":
    init_path = r"D:\PYTHON Practice\Consolidate function folder"
    file_name_keyword = "census"
    sheet_name_keyword = "population"
    destination_path = r"D:\PYTHON Practice\Consolidate function folder - result"
    destination_file_name = "Consolidation results.xlsx"
    skip_folders = ["Fold AL, VV"]

    logging.basicConfig(level=logging.CRITICAL)
    consolidate_excel_data(init_path, file_name_keyword, sheet_name_keyword, destination_path, destination_file_name,
                           skip_folders)

