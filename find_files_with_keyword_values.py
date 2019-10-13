import os
import openpyxl
import logging
from openpyxl.utils.cell import get_column_letter
import datetime
from tkinker_gui import get_user_parameters_with_gui


def find_files_with_keyword_values(search_keywords: list, working_directory: str = None, exclusion_folder: str = None,
                                   file_name_keyword: str = None, search_sheet: str = None):
    """
    Finds excel files that contain certain keyword values.\n
    It iterates through all files from the specified directory including subfolders.\n
    It can run search in all files as well as in file with specific keyword in file name.
    It can also search values in all spreadsheets as well as in specific spreadsheet.\n
    :param search_keywords: list
    :param working_directory: absolute path, optional
    :param exclusion_folder: str, optional
    :param file_name_keyword: str, optional (if not passed, function processes all excel files)
    :param search_sheet: str, optional (if not passed, function processes all worksheets)
    :return:
    """

    def search_keyword_in_sheets(wb: openpyxl, sheet_name: str, search_keywords: list, abs_path: str):
        # creating generator for lazy cell iteration (memory efficient on big volumes)
        def get_cell(wb, sheet_name):
            for row in wb[sheet_name].iter_rows():
                for cell in row:
                    yield cell

        for cell_keyword in search_keywords:
            for cell in get_cell(wb, sheet_name):
                if cell_keyword.lower() in str(cell.value).lower():
                    print(f"    >>> FOUND KEYWORD '{cell_keyword}': cell value '{cell.value}', sheet '{sheet_name}'"
                          f", coordinates '{get_column_letter(cell.column)}:{cell.row}', file [{abs_path}]")

    def search_keyword_in_file(abs_path: str, search_sheet: str):
        # open file
        try:
            wb = openpyxl.load_workbook(abs_path)
        except FileNotFoundError as err:
            print(f"{err} for file '{file}'")
        else:
            # search keyword in excel values for specific sheet or for all sheets
            if search_sheet:
                matched_sheets = [sheet for sheet in wb.sheetnames if search_sheet.lower() in sheet.lower()]
                if matched_sheets:
                    for matched_sheet in matched_sheets:
                        search_keyword_in_sheets(wb, matched_sheet, search_keywords, abs_path)
            else:
                for active_sheet in wb.sheetnames:
                    search_keyword_in_sheets(wb, active_sheet, search_keywords, abs_path)

    start = datetime.datetime.now()

    # switch to working_directory if passed as param
    if working_directory:
        os.chdir(working_directory)
    
    # walk the directory tree
    print("Process started...")
    for active_folder, subfolders, files in os.walk('.'):

        # skip exclusion folder from search if passed as param
        if exclusion_folder:
            if exclusion_folder in active_folder.split(os.sep):
                print(f"Excluded '{active_folder}' from search as it contains exclusion folder '{exclusion_folder}'")
                continue

        for file in files:
            file_path = os.path.join(active_folder, file)
            abs_path = os.path.abspath(file_path)  # get abs path to open file

            # search for excel files only
            if file.endswith('xlsx'):
                logging.debug(f"Processing {abs_path} ...")

                # search for files with keyword in name
                if file_name_keyword:
                    if file_name_keyword.lower() in file.lower():
                        search_keyword_in_file(abs_path, search_sheet)
                else:
                    search_keyword_in_file(abs_path, search_sheet)

    finish = datetime.datetime.now()
    time_spent = finish - start
    print(f"Process finished. Time spent {time_spent}")


if __name__ == "__main__":

    # default params
    params_input = {"file_name_keyword": [0, "file"],
                    "search_keywords": [0, "Castle, Final Fantasy"],
                    "search_sheet": [0, "sheet"],
                    "working_directory": [1, r"D:\PYTHON Practice\Top Secret"],
                    "exclusion_folder": [0, "Top Secret"],
                    }
    # getting output from user with GUI
    params_output = get_user_parameters_with_gui(params_input)

    # parsing user output params
    file_keyword = params_output["file_name_keyword"]
    cell_keywords = [value.strip() for value in params_output["search_keywords"].split(",")]
    sheet_name = params_output["search_sheet"]
    directory = params_output["working_directory"]
    exclusion_folder_name = params_output["exclusion_folder"]

    logging.basicConfig(level=logging.CRITICAL, format=' %(asctime)s - %(levelname)s - %(message)s')
    find_files_with_keyword_values(search_keywords=cell_keywords, working_directory=directory,
                                   exclusion_folder=None, file_name_keyword=file_keyword,
                                   search_sheet=sheet_name)

