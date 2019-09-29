import openpyxl
import os
import datetime
from openpyxl.utils.cell import get_column_letter
from operator import itemgetter


def generate_schedule(file_name, schedule_data, iteration_days, iteration_number, date_format,
                      tasks_execution_sheet_name, schedule_sheet_name, cwd=None):
    """
    Generates and records task schedule into Excel workbook based on the task-date input.\n
    To do that it creates all_tasks dictionary with all tasks, their last run date and next iterations run dates.
    Then it records preliminary data into Excel workbook, creates new task_schedule list, sorts tasks by date
    and records final schedule into separate spreadsheet.\n
    As a final result it creates 2 spreadsheets - with tasks and iteration dates and with final schedule sorted by date.\n
    :param file_name: .xlsx file
    :param schedule_data: dictionary
    :param iteration_days: datetime
    :param iteration_number: int
    :param date_format: str
    :param tasks_execution_sheet_name: str
    :param schedule_sheet_name: str
    :param cwd: valid path
    :return:
    """

    # switch to cwd if passed as an argument
    if cwd:
        if not os.path.isdir(cwd):
            os.mkdir(cwd)
        os.chdir(cwd)

    sheet_name = schedule_data["sheet_name"]
    tasks_col = schedule_data["tasks_col"]
    date_col = schedule_data["date_col"]

    # create list with task and last run date
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]

    all_tasks = {}
    max_row = sheet.max_row
    for i in range(2, max_row+1):  # skip header
        task = sheet.cell(row=i, column=tasks_col).value
        date = sheet.cell(row=i, column=date_col).value
        if task not in all_tasks.keys():  # record unique tasks
            all_tasks.setdefault(task, [date])
        if date > all_tasks[task][0]:  # overwrite with most recent date
            all_tasks[task] = [date]

    # calculate next execution day
    for task in all_tasks.keys():
        for i in range(iteration_number):
            date = all_tasks[task][i]
            next_date = date + iteration_days
            while next_date.weekday() in (0, 5, 6):  # do not schedule for Monday and weekend
                next_date += datetime.timedelta(1)
            all_tasks[task].append(next_date)

    # (re)create new sheet for task-dates results
    if tasks_execution_sheet_name in wb.sheetnames:
        del wb[tasks_execution_sheet_name]
    wb.create_sheet(tasks_execution_sheet_name)
    print(f"Recreated '{tasks_execution_sheet_name}' sheet")
    result_sheet = wb[tasks_execution_sheet_name]

    # populate sheet with all tasks and their dates
    row = 1
    for task in all_tasks.keys():
        result_sheet[f"A{row}"] = task
        for col in range(1, len(all_tasks[task])+1):
            date_cell = result_sheet[f"{get_column_letter(col+1)}{row}"]
            date_cell.value = all_tasks[task][col-1]  # record value
            date_cell.number_format = date_format  # set date format
        row += 1

    # set column width
    tasks_max_len = max(len(task) for task in all_tasks.keys())*1.5  # 1.5 - coefficient for extra space
    date_max_len = max(len(str(date)) for dates in all_tasks.values() for date in dates)  # list compr. with nested loop
    result_sheet.column_dimensions["A"].width = tasks_max_len
    for i in range(2, iteration_number+3):
        result_sheet.column_dimensions[get_column_letter(i)].width = date_max_len

    print(f"    Recorded results into '{tasks_execution_sheet_name}' sheet")

    # prepare task schedule and sort records
    tasks_schedule = []
    for i in range(iteration_number+1):
        for task in all_tasks.keys():
            tasks_schedule.append([task, all_tasks[task][i]])

    tasks_schedule = sorted(tasks_schedule, key=itemgetter(1))

    # (re)create new sheet for schedule
    if schedule_sheet_name in wb.sheetnames:
        del wb[schedule_sheet_name]
    wb.create_sheet(schedule_sheet_name)
    print(f"Recreated '{schedule_sheet_name}' sheet")
    result_sheet = wb[schedule_sheet_name]

    # populate sheet with schedule data
    for row in range(len(tasks_schedule)):
        for col in range(len(tasks_schedule[row])):
            result_sheet[f"{get_column_letter(col+1)}{row+1}"] = tasks_schedule[row][col]
    print(f"    Recorded schedule into '{schedule_sheet_name}' sheet\nProcess finished")

    wb.save(file_name)


if __name__ == "__main__":

    working_directory = r"D:\Practice Python\Tasks schedule generator"
    file = "Tasks_schedule.xlsx"
    file_data = {"sheet_name": "Schedule", "tasks_col": 1, "date_col": 2}
    iterations_days = datetime.timedelta(90)
    iterations_number = 4
    date_formatting = "MM/DD/YYYY"
    tasks_execution_sheet = "Tasks_and_execution_dates"
    schedule_sheet = "Tasks_schedule"

    generate_schedule(file, file_data, iterations_days, iterations_number, date_formatting,
                      tasks_execution_sheet, schedule_sheet)

